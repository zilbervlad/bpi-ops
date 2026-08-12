from collections import defaultdict
from datetime import datetime, time, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo

from flask import Blueprint, render_template, request, session, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.auth.routes import login_required, role_required
from app.models import CashLog, Store

cash_review_bp = Blueprint("cash_review", __name__, url_prefix="/cash-review")

APP_TZ = ZoneInfo("America/New_York")
BUSINESS_DAY_CUTOFF = time(5, 0)
MORNING_COUNT_DUE = time(10, 30)
MIDSHIFT_COUNT_DUE = time(15, 0)
NIGHT_COUNT_DUE = time(3, 0)
CASH_TOLERANCE = 0.01


def now_et():
    return datetime.now(APP_TZ)


def get_visible_stores():
    role = session.get("user_role")
    user_area = session.get("user_area")
    user_store = session.get("user_store")

    if role == "admin":
        return Store.query.filter_by(is_active=True).order_by(Store.store_number.asc()).all()

    if role == "supervisor":
        return Store.query.filter_by(
            area_name=user_area,
            is_active=True,
        ).order_by(Store.store_number.asc()).all()

    if role == "general_manager" and user_store:
        return Store.query.filter_by(
            store_number=user_store,
            is_active=True,
        ).order_by(Store.store_number.asc()).all()

    return []


def current_cash_review_date():
    current = now_et()
    if current.time() < BUSINESS_DAY_CUTOFF:
        return (current - timedelta(days=1)).date()
    return current.date()


def build_shift_due_status(dashboard_date):
    current = now_et()
    active_ops_date = current_cash_review_date()

    if dashboard_date < active_ops_date:
        return {
            "opening": True,
            "midshift": True,
            "closing": True,
        }

    if dashboard_date > active_ops_date:
        return {
            "opening": False,
            "midshift": False,
            "closing": False,
        }

    # Between midnight and the 5 AM ops cutoff we are still reviewing the
    # previous business day. Opening and midshift are already due; the night
    # count becomes due once the stores reach the late-night close window.
    if current.time() < BUSINESS_DAY_CUTOFF:
        return {
            "opening": True,
            "midshift": True,
            "closing": current.time() >= NIGHT_COUNT_DUE,
        }

    return {
        "opening": current.time() >= MORNING_COUNT_DUE,
        "midshift": current.time() >= MIDSHIFT_COUNT_DUE,
        "closing": False,
    }


def build_closing_to_opening_diffs(logs):
    by_store = defaultdict(list)

    for log in logs:
        by_store[log.store_number].append(log)

    diff_rows = []

    for store_number, store_logs in by_store.items():
        ordered = sorted(
            store_logs,
            key=lambda x: (
                x.log_date,
                x.created_at or datetime.min,
                x.id or 0,
            ),
        )

        for index, current_log in enumerate(ordered):
            if current_log.shift_type != "closing":
                continue

            next_opening = None
            for future_log in ordered[index + 1:]:
                if future_log.shift_type == "opening":
                    next_opening = future_log
                    break

            if not next_opening:
                continue

            closing_total = current_log.total_cash or 0
            opening_total = next_opening.total_cash or 0
            diff_amount = opening_total - closing_total

            diff_rows.append({
                "store_number": store_number,
                "closing_date": current_log.log_date,
                "opening_date": next_opening.log_date,
                "closing_total": closing_total,
                "opening_total": opening_total,
                "difference": diff_amount,
                "closing_manager": current_log.manager_name,
                "opening_manager": next_opening.manager_name,
            })

    diff_rows.sort(
        key=lambda row: (
            abs(row["difference"] or 0),
            row["closing_date"],
            row["store_number"],
        ),
        reverse=True,
    )
    return diff_rows


def build_attention_rows(
    missed_morning_stores,
    missed_midshift_stores,
    missed_night_stores,
    midshift_exceptions,
    closing_opening_diffs,
):
    attention = {}

    def row_for(store_number):
        store_number = str(store_number)
        if store_number not in attention:
            attention[store_number] = {
                "store_number": store_number,
                "missing": [],
                "midshift_variance": None,
                "midshift_manager": None,
                "handoff_difference": None,
                "handoff_manager": None,
                "issue_count": 0,
                "largest_amount": 0.0,
            }
        return attention[store_number]

    for store_number in missed_morning_stores:
        row_for(store_number)["missing"].append("Morning")

    for store_number in missed_midshift_stores:
        row_for(store_number)["missing"].append("Midshift")

    for store_number in missed_night_stores:
        row_for(store_number)["missing"].append("Night")

    for log in midshift_exceptions:
        row = row_for(log.store_number)
        variance = log.cash_over_short or 0
        current_value = row["midshift_variance"]

        if current_value is None or abs(variance) > abs(current_value):
            row["midshift_variance"] = variance
            row["midshift_manager"] = log.manager_name

    for handoff in closing_opening_diffs:
        row = row_for(handoff["store_number"])
        difference = handoff["difference"] or 0
        current_value = row["handoff_difference"]

        if current_value is None or abs(difference) > abs(current_value):
            row["handoff_difference"] = difference
            row["handoff_manager"] = (
                handoff.get("opening_manager")
                or handoff.get("closing_manager")
            )

    rows = []
    for row in attention.values():
        row["issue_count"] = (
            len(row["missing"])
            + (1 if row["midshift_variance"] is not None else 0)
            + (1 if row["handoff_difference"] is not None else 0)
        )
        row["largest_amount"] = max(
            abs(row["midshift_variance"] or 0),
            abs(row["handoff_difference"] or 0),
        )
        row["manager_name"] = (
            row["midshift_manager"]
            or row["handoff_manager"]
            or ""
        )
        rows.append(row)

    rows.sort(
        key=lambda row: (
            row["issue_count"],
            row["largest_amount"],
            row["store_number"],
        ),
        reverse=True,
    )
    return rows


def build_cash_review_payload():
    visible_stores = get_visible_stores()
    visible_store_numbers = {store.store_number for store in visible_stores}

    store_filter = (request.args.get("store") or "").strip()
    shift_filter = (request.args.get("shift") or "").strip()
    date_filter = (request.args.get("date") or "").strip()

    today = current_cash_review_date()
    selected_date = None

    if date_filter:
        try:
            selected_date = datetime.strptime(date_filter, "%Y-%m-%d").date()
        except ValueError:
            selected_date = None
            date_filter = ""

    dashboard_date = selected_date or today

    scope_stores = visible_stores
    if store_filter:
        scope_stores = [
            store for store in visible_stores
            if str(store.store_number) == str(store_filter)
        ]

    scope_store_numbers = {store.store_number for store in scope_stores}

    dashboard_logs = CashLog.query.filter(
        CashLog.store_number.in_(scope_store_numbers),
        CashLog.log_date == dashboard_date,
    ).all() if scope_store_numbers else []

    logs = dashboard_logs
    if shift_filter:
        logs = [
            log for log in logs
            if log.shift_type == shift_filter
        ]

    logs = sorted(
        logs,
        key=lambda log: (
            log.created_at or datetime.min,
            log.id or 0,
        ),
        reverse=True,
    )[:100]

    stores_by_shift = {
        "opening": set(),
        "midshift": set(),
        "closing": set(),
    }

    for log in dashboard_logs:
        if log.shift_type in stores_by_shift:
            stores_by_shift[log.shift_type].add(log.store_number)

    shift_due = build_shift_due_status(dashboard_date)

    raw_missing_morning = [
        store.store_number for store in scope_stores
        if store.store_number not in stores_by_shift["opening"]
    ]
    raw_missing_midshift = [
        store.store_number for store in scope_stores
        if store.store_number not in stores_by_shift["midshift"]
    ]
    raw_missing_night = [
        store.store_number for store in scope_stores
        if store.store_number not in stores_by_shift["closing"]
    ]

    missed_morning_stores = raw_missing_morning if shift_due["opening"] else []
    missed_midshift_stores = raw_missing_midshift if shift_due["midshift"] else []
    missed_night_stores = raw_missing_night if shift_due["closing"] else []

    midshift_all_logs = [
        log for log in dashboard_logs
        if log.shift_type == "midshift"
    ]

    midshift_exceptions = [
        log for log in midshift_all_logs
        if log.cash_over_short is not None
        and abs(log.cash_over_short) >= CASH_TOLERANCE
    ]
    midshift_exceptions.sort(
        key=lambda log: (
            abs(log.cash_over_short or 0),
            log.store_number,
        ),
        reverse=True,
    )

    diff_base_query = CashLog.query.filter(
        CashLog.store_number.in_(scope_store_numbers),
        CashLog.log_date >= dashboard_date - timedelta(days=1),
        CashLog.log_date <= dashboard_date,
    ) if scope_store_numbers else None

    diff_logs = diff_base_query.all() if diff_base_query is not None else []
    closing_opening_pairs = build_closing_to_opening_diffs(diff_logs)
    closing_opening_diffs = [
        row for row in closing_opening_pairs
        if abs(row["difference"] or 0) >= CASH_TOLERANCE
    ]

    attention_rows = build_attention_rows(
        missed_morning_stores=missed_morning_stores,
        missed_midshift_stores=missed_midshift_stores,
        missed_night_stores=missed_night_stores,
        midshift_exceptions=midshift_exceptions,
        closing_opening_diffs=closing_opening_diffs,
    )

    total_abs_variance = round(
        sum(abs(log.cash_over_short or 0) for log in midshift_exceptions),
        2,
    )
    total_abs_handoff = round(
        sum(abs(row["difference"] or 0) for row in closing_opening_diffs),
        2,
    )

    largest_candidates = []
    for log in midshift_exceptions:
        largest_candidates.append({
            "amount": abs(log.cash_over_short or 0),
            "signed_amount": log.cash_over_short or 0,
            "store_number": log.store_number,
            "type": "Midshift",
        })

    for row in closing_opening_diffs:
        largest_candidates.append({
            "amount": abs(row["difference"] or 0),
            "signed_amount": row["difference"] or 0,
            "store_number": row["store_number"],
            "type": "Handoff",
        })

    largest_issue = max(
        largest_candidates,
        key=lambda row: row["amount"],
        default=None,
    )

    summary = {
        "activity_count": len(dashboard_logs),
        "log_count": len(logs),
        "stores_in_scope": len(scope_stores),
        "attention_store_count": len(attention_rows),
        "missing_total": (
            len(missed_morning_stores)
            + len(missed_midshift_stores)
            + len(missed_night_stores)
        ),
        "missed_night_count": len(missed_night_stores),
        "missed_morning_count": len(missed_morning_stores),
        "missed_midshift_count": len(missed_midshift_stores),
        "morning_complete_count": len(stores_by_shift["opening"]),
        "midshift_complete_count": len(stores_by_shift["midshift"]),
        "night_complete_count": len(stores_by_shift["closing"]),
        "morning_due": shift_due["opening"],
        "midshift_due": shift_due["midshift"],
        "night_due": shift_due["closing"],
        "midshift_logged_count": len(midshift_all_logs),
        "midshift_count": len(midshift_exceptions),
        "diff_pair_count": len(closing_opening_diffs),
        "handoff_pair_count": len(closing_opening_pairs),
        "total_abs_variance": total_abs_variance,
        "total_abs_handoff": total_abs_handoff,
        "largest_issue": largest_issue,
    }

    return {
        "stores": visible_stores,
        "logs": logs,
        "midshift_logs": midshift_exceptions,
        "closing_opening_diffs": closing_opening_diffs,
        "attention_rows": attention_rows,
        "missed_night_stores": missed_night_stores,
        "missed_morning_stores": missed_morning_stores,
        "missed_midshift_stores": missed_midshift_stores,
        "store_filter": store_filter,
        "shift_filter": shift_filter,
        "date_filter": date_filter,
        "dashboard_date": dashboard_date,
        "summary": summary,
    }


def autosize_worksheet_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            except Exception:
                pass

        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            40,
        )


def style_header_row(worksheet, row_number=1):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[row_number]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_cash_review_excel(payload):
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(["Metric", "Value"])

    summary_rows = [
        ("Selected Store", payload["store_filter"] or "All"),
        ("Selected Shift", payload["shift_filter"] or "All"),
        ("Selected Date", payload["dashboard_date"].isoformat()),
        ("Stores In Scope", payload["summary"]["stores_in_scope"]),
        ("Stores Needing Review", payload["summary"]["attention_store_count"]),
        ("Missing Required Counts", payload["summary"]["missing_total"]),
        ("Midshift Variance Exceptions", payload["summary"]["midshift_count"]),
        ("Absolute Midshift Variance", payload["summary"]["total_abs_variance"]),
        ("Handoff Discrepancies", payload["summary"]["diff_pair_count"]),
        ("Absolute Handoff Difference", payload["summary"]["total_abs_handoff"]),
    ]

    for row in summary_rows:
        summary_ws.append(row)

    style_header_row(summary_ws)
    autosize_worksheet_columns(summary_ws)

    attention_ws = wb.create_sheet(title="Needs Attention")
    attention_ws.append([
        "Store",
        "Missing Counts",
        "Midshift Over Short",
        "Handoff Difference",
        "Manager",
    ])

    for row in payload["attention_rows"]:
        attention_ws.append([
            row["store_number"],
            ", ".join(row["missing"]),
            row["midshift_variance"] if row["midshift_variance"] is not None else "",
            row["handoff_difference"] if row["handoff_difference"] is not None else "",
            row["manager_name"],
        ])

    style_header_row(attention_ws)
    autosize_worksheet_columns(attention_ws)

    logs_ws = wb.create_sheet(title="Recent Cash Logs")
    logs_ws.append([
        "Store",
        "Date",
        "Shift",
        "Back Till",
        "Front Till",
        "Driver Banks",
        "Total Cash",
        "Amount To Account For",
        "Cash Over / Short",
        "Manager",
    ])

    for log in payload["logs"]:
        logs_ws.append([
            log.store_number,
            log.log_date.strftime("%Y-%m-%d") if log.log_date else "",
            log.shift_type.title() if log.shift_type else "",
            log.back_till if log.back_till is not None else "",
            log.front_till if log.front_till is not None else "",
            log.driver_banks if log.driver_banks is not None else "",
            log.total_cash if log.total_cash is not None else "",
            log.amount_to_account_for if log.amount_to_account_for is not None else "",
            log.cash_over_short if log.cash_over_short is not None else "",
            log.manager_name or "",
        ])

    style_header_row(logs_ws)
    autosize_worksheet_columns(logs_ws)

    midshift_ws = wb.create_sheet(title="Midshift Exceptions")
    midshift_ws.append([
        "Store",
        "Date",
        "Total Cash",
        "Amount To Account For",
        "Cash Over / Short",
        "Manager",
    ])

    for log in payload["midshift_logs"]:
        midshift_ws.append([
            log.store_number,
            log.log_date.strftime("%Y-%m-%d") if log.log_date else "",
            log.total_cash if log.total_cash is not None else "",
            log.amount_to_account_for if log.amount_to_account_for is not None else "",
            log.cash_over_short if log.cash_over_short is not None else "",
            log.manager_name or "",
        ])

    style_header_row(midshift_ws)
    autosize_worksheet_columns(midshift_ws)

    diff_ws = wb.create_sheet(title="Handoff Discrepancies")
    diff_ws.append([
        "Store",
        "Closing Date",
        "Closing Total",
        "Closing Manager",
        "Opening Date",
        "Opening Total",
        "Opening Manager",
        "Difference",
    ])

    for row in payload["closing_opening_diffs"]:
        diff_ws.append([
            row["store_number"],
            row["closing_date"].strftime("%Y-%m-%d") if row["closing_date"] else "",
            row["closing_total"],
            row["closing_manager"] or "",
            row["opening_date"].strftime("%Y-%m-%d") if row["opening_date"] else "",
            row["opening_total"],
            row["opening_manager"] or "",
            row["difference"],
        ])

    style_header_row(diff_ws)
    autosize_worksheet_columns(diff_ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@cash_review_bp.route("/", methods=["GET"])
@login_required
@role_required("admin", "supervisor", "general_manager")
def index():
    payload = build_cash_review_payload()

    return render_template(
        "cash_review.html",
        stores=payload["stores"],
        logs=payload["logs"],
        midshift_logs=payload["midshift_logs"],
        closing_opening_diffs=payload["closing_opening_diffs"],
        attention_rows=payload["attention_rows"],
        missed_night_stores=payload["missed_night_stores"],
        missed_morning_stores=payload["missed_morning_stores"],
        missed_midshift_stores=payload["missed_midshift_stores"],
        store_filter=payload["store_filter"],
        shift_filter=payload["shift_filter"],
        date_filter=payload["date_filter"],
        dashboard_date=payload["dashboard_date"],
        summary=payload["summary"],
    )


@cash_review_bp.route("/export/excel", methods=["GET"])
@login_required
@role_required("admin", "supervisor", "general_manager")
def export_excel():
    payload = build_cash_review_payload()
    workbook_stream = create_cash_review_excel(payload)

    filename_parts = ["cash_review"]
    if payload["store_filter"]:
        filename_parts.append(payload["store_filter"])
    if payload["shift_filter"]:
        filename_parts.append(payload["shift_filter"])
    filename_parts.append(payload["dashboard_date"].isoformat())

    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        workbook_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )