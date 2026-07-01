from datetime import datetime, date, time, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.auth.routes import login_required, role_required
from app.extensions import db
from app.models import MaintenanceTicket, MaintenanceTimeCard, Store, User

maintenance_bp = Blueprint("maintenance", __name__, url_prefix="/maintenance")

APP_TZ = ZoneInfo("America/New_York")
UTC_TZ = ZoneInfo("UTC")


def now_utc_naive():
    return datetime.utcnow()


def now_et():
    return datetime.now(APP_TZ)


def today_et_local():
    return now_et().date()


def utc_naive_to_et(dt):
    if not dt:
        return None
    return dt.replace(tzinfo=UTC_TZ).astimezone(APP_TZ)


def et_local_to_utc_naive(dt):
    if not dt:
        return None
    return dt.replace(tzinfo=APP_TZ).astimezone(UTC_TZ).replace(tzinfo=None)


def parse_datetime_local(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        local_dt = datetime.strptime(value, "%Y-%m-%dT%H:%M")
    except ValueError:
        return None
    return et_local_to_utc_naive(local_dt)


def format_dt_et(dt):
    local_dt = utc_naive_to_et(dt)
    if not local_dt:
        return ""
    return local_dt.strftime("%-I:%M %p")


def format_dt_local_input(dt):
    local_dt = utc_naive_to_et(dt)
    if not local_dt:
        return ""
    return local_dt.strftime("%Y-%m-%dT%H:%M")


def hours_between(clock_in_at, clock_out_at):
    if not clock_in_at or not clock_out_at:
        return None
    if clock_out_at <= clock_in_at:
        return None
    return round((clock_out_at - clock_in_at).total_seconds() / 3600.0, 2)


def monday_for_date(value):
    return value - timedelta(days=value.weekday())


def sunday_for_date(value):
    return monday_for_date(value) + timedelta(days=6)


def get_current_user_record():
    user_id = session.get("user_id")
    if user_id:
        user = User.query.get(user_id)
        if user:
            return user

    username = session.get("username") or session.get("user_username")
    if username:
        user = User.query.filter_by(username=username).first()
        if user:
            return user

    user_name = session.get("user_name")
    if user_name:
        return User.query.filter_by(name=user_name).first()

    return None


def get_or_create_time_card(user_id, work_date):
    card = MaintenanceTimeCard.query.filter_by(
        user_id=user_id,
        work_date=work_date
    ).first()

    if card:
        return card

    card = MaintenanceTimeCard(
        user_id=user_id,
        work_date=work_date,
    )
    db.session.add(card)
    db.session.flush()
    return card


def get_maintenance_users():
    return (
        User.query
        .filter(User.role == "maintenance", User.is_active == True)
        .order_by(User.name.asc())
        .all()
    )


def build_time_card_report(start_date, end_date, tech_id=None):
    maintenance_users = get_maintenance_users()
    user_ids = [u.id for u in maintenance_users]

    if tech_id:
        user_ids = [tech_id]

    full_start = monday_for_date(start_date)
    full_end = sunday_for_date(end_date)

    full_cards = (
        MaintenanceTimeCard.query
        .filter(
            MaintenanceTimeCard.user_id.in_(user_ids),
            MaintenanceTimeCard.work_date >= full_start,
            MaintenanceTimeCard.work_date <= full_end,
        )
        .order_by(
            MaintenanceTimeCard.user_id.asc(),
            MaintenanceTimeCard.work_date.asc(),
            MaintenanceTimeCard.clock_in_at.asc(),
            MaintenanceTimeCard.id.asc(),
        )
        .all()
    )

    full_cards_by_user_week = {}
    for card in full_cards:
        week_start = monday_for_date(card.work_date)
        full_cards_by_user_week.setdefault((card.user_id, week_start), []).append(card)

    allocations = {}
    for key, cards in full_cards_by_user_week.items():
        running_hours = 0.0

        for card in cards:
            total = hours_between(card.clock_in_at, card.clock_out_at)

            if total is None:
                allocations[card.id] = {
                    "regular": None,
                    "ot": None,
                    "total": None,
                    "missing": True,
                }
                continue

            regular_remaining = max(40.0 - running_hours, 0.0)
            regular = min(total, regular_remaining)
            ot = max(total - regular, 0.0)

            allocations[card.id] = {
                "regular": round(regular, 2),
                "ot": round(ot, 2),
                "total": round(total, 2),
                "missing": False,
            }

            running_hours += total

    display_cards = [
        card for card in full_cards
        if start_date <= card.work_date <= end_date
    ]

    users_by_id = {u.id: u for u in maintenance_users}
    summaries = {}

    for user in maintenance_users:
        if tech_id and user.id != tech_id:
            continue
        summaries[user.id] = {
            "user": user,
            "regular": 0.0,
            "ot": 0.0,
            "total": 0.0,
            "missing": 0,
            "days_worked": 0,
        }

    rows = []
    for card in display_cards:
        user = users_by_id.get(card.user_id)
        if not user:
            continue

        allocation = allocations.get(card.id, {})
        missing = allocation.get("missing", True)

        if not card.clock_in_at or not card.clock_out_at:
            missing = True

        row = {
            "card": card,
            "user": user,
            "date": card.work_date,
            "day": card.work_date.strftime("%a"),
            "clock_in": format_dt_et(card.clock_in_at) if card.clock_in_at else "Missing",
            "clock_out": format_dt_et(card.clock_out_at) if card.clock_out_at else "Missing",
            "clock_in_input": format_dt_local_input(card.clock_in_at),
            "clock_out_input": format_dt_local_input(card.clock_out_at),
            "regular": allocation.get("regular"),
            "ot": allocation.get("ot"),
            "total": allocation.get("total"),
            "missing": missing,
            "notes": card.notes or "",
            "is_edited": bool(getattr(card, "is_edited", False)),
            "edited_at": format_dt_et(card.edited_at) if getattr(card, "edited_at", None) else "",
            "edited_by": card.edited_by.name if getattr(card, "edited_by", None) else "",
        }
        rows.append(row)

        summary = summaries.setdefault(card.user_id, {
            "user": user,
            "regular": 0.0,
            "ot": 0.0,
            "total": 0.0,
            "missing": 0,
            "days_worked": 0,
        })

        if missing:
            summary["missing"] += 1
        else:
            summary["regular"] += allocation.get("regular") or 0.0
            summary["ot"] += allocation.get("ot") or 0.0
            summary["total"] += allocation.get("total") or 0.0
            summary["days_worked"] += 1

    for summary in summaries.values():
        summary["regular"] = round(summary["regular"], 2)
        summary["ot"] = round(summary["ot"], 2)
        summary["total"] = round(summary["total"], 2)

    total_summary = {
        "regular": round(sum(s["regular"] for s in summaries.values()), 2),
        "ot": round(sum(s["ot"] for s in summaries.values()), 2),
        "total": round(sum(s["total"] for s in summaries.values()), 2),
        "missing": sum(s["missing"] for s in summaries.values()),
    }

    return {
        "maintenance_users": maintenance_users,
        "summaries": list(summaries.values()),
        "rows": rows,
        "total_summary": total_summary,
    }


def create_time_cards_pdf(start_date, end_date, tech_id=None):
    report = build_time_card_report(start_date, end_date, tech_id=tech_id)

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=0.35 * inch,
        leftMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TimeCardTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        spaceAfter=8,
    )
    small_style = ParagraphStyle(
        "TimeCardSmall",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
    )

    elements = []

    elements.append(Paragraph("Boston Pie Inc. - Maintenance Time Card Report", title_style))
    elements.append(Paragraph(
        f"Date Range: {start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')} | Generated: {now_et().strftime('%m/%d/%Y %-I:%M %p')} ET",
        small_style,
    ))
    elements.append(Spacer(1, 0.12 * inch))

    summary_data = [["Tech", "Days Worked", "Regular Hours", "OT Hours", "Total Hours", "Missing Punches"]]
    for summary in report["summaries"]:
        summary_data.append([
            summary["user"].name,
            str(summary["days_worked"]),
            f'{summary["regular"]:.2f}',
            f'{summary["ot"]:.2f}',
            f'{summary["total"]:.2f}',
            str(summary["missing"]),
        ])

    summary_data.append([
        "TOTAL",
        "",
        f'{report["total_summary"]["regular"]:.2f}',
        f'{report["total_summary"]["ot"]:.2f}',
        f'{report["total_summary"]["total"]:.2f}',
        str(report["total_summary"]["missing"]),
    ])

    summary_table = Table(summary_data, repeatRows=1, colWidths=[1.8*inch, 1.0*inch, 1.2*inch, 1.0*inch, 1.1*inch, 1.2*inch])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D9EAF7")),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#A6A6A6")),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.18 * inch))

    detail_data = [["Date", "Day", "Tech", "Clock In", "Clock Out", "Reg", "OT", "Total", "Edited", "Notes"]]
    for row in report["rows"]:
        detail_data.append([
            row["date"].strftime("%m/%d/%y"),
            row["day"],
            row["user"].name,
            row["clock_in"],
            row["clock_out"],
            "" if row["regular"] is None else f'{row["regular"]:.2f}',
            "" if row["ot"] is None else f'{row["ot"]:.2f}',
            "" if row["total"] is None else f'{row["total"]:.2f}',
            "Yes" if row.get("is_edited") else "",
            Paragraph((row["notes"] or "-")[:180], small_style),
        ])

    if len(detail_data) == 1:
        detail_data.append(["", "", "No time cards in selected range.", "", "", "", "", "", "", ""])

    detail_table = Table(
        detail_data,
        repeatRows=1,
        colWidths=[0.70*inch, 0.40*inch, 1.10*inch, 0.78*inch, 0.78*inch, 0.48*inch, 0.48*inch, 0.55*inch, 0.55*inch, 2.65*inch],
    )
    detail_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BFBFBF")),
        ("ALIGN", (0, 1), (7, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(detail_table)

    doc.build(elements)
    output.seek(0)
    return output


def get_current_role():
    return (session.get("user_role") or "").strip()


def get_visible_stores():
    role = session.get("user_role")
    user_area = session.get("user_area")
    user_store = session.get("user_store")

    if role in ["admin", "maintenance"]:
        return Store.query.filter_by(is_active=True).order_by(Store.store_number.asc()).all()

    if role == "supervisor":
        return Store.query.filter_by(
            area_name=user_area,
            is_active=True
        ).order_by(Store.store_number.asc()).all()

    if role == "manager":
        return Store.query.filter_by(
            store_number=user_store,
            is_active=True
        ).order_by(Store.store_number.asc()).all()

    return []


def get_visible_store_numbers():
    return {store.store_number for store in get_visible_stores()}


def split_lines_to_tasks(text: str):
    return [line.strip() for line in (text or "").splitlines() if line.strip()]


def get_filtered_tickets():
    visible_stores = get_visible_stores()
    visible_store_numbers = {store.store_number for store in visible_stores}

    status_filter = request.args.get("status", "").strip()
    store_filter = request.args.get("store", "").strip()

    tickets = MaintenanceTicket.query.order_by(
        MaintenanceTicket.created_at.asc(),
        MaintenanceTicket.id.asc()
    ).all()

    tickets = [t for t in tickets if t.store_number in visible_store_numbers]

    if status_filter:
        tickets = [t for t in tickets if t.status == status_filter]

    if store_filter:
        tickets = [t for t in tickets if t.store_number == store_filter]

    return tickets, visible_stores, status_filter, store_filter


def autosize_worksheet_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            except Exception:
                pass

        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


def style_header_row(worksheet, row_number=1):
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[row_number]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_maintenance_excel(tickets, status_filter, store_filter):
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["Selected Store", store_filter or "All"])
    summary_ws.append(["Selected Status", status_filter or "All"])
    summary_ws.append(["Total Tickets", len(tickets)])

    style_header_row(summary_ws)
    autosize_worksheet_columns(summary_ws)

    tickets_ws = wb.create_sheet(title="Maintenance Tickets")
    tickets_ws.append([
        "Store",
        "Title",
        "Details",
        "Status",
        "Source Type",
        "Created At",
        "SVR Report ID",
    ])

    for ticket in tickets:
        created_at_display = (
            ticket.created_at.strftime("%Y-%m-%d %I:%M %p")
            if ticket.created_at else ""
        )

        tickets_ws.append([
            ticket.store_number,
            ticket.title or "",
            ticket.details or "",
            ticket.status.replace("_", " ").title() if ticket.status else "",
            (ticket.source_type or "").upper(),
            created_at_display,
            ticket.svr_report_id if ticket.svr_report_id is not None else "",
        ])

    style_header_row(tickets_ws)
    autosize_worksheet_columns(tickets_ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output



def parse_optional_date(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_optional_time(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%H:%M").time()
    except ValueError:
        return None


def parse_optional_int(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        parsed = int(value)
        return parsed if parsed >= 0 else None
    except ValueError:
        return None


def normalize_priority(value):
    value = (value or "normal").strip().lower()
    valid_priorities = {"low", "normal", "high", "urgent"}
    return value if value in valid_priorities else "normal"


def format_maintenance_time(ticket):
    if not ticket or not getattr(ticket, "scheduled_time", None):
        return "Any Time"

    try:
        return ticket.scheduled_time.strftime("%I:%M %p").lstrip("0")
    except Exception:
        return "Any Time"


def build_calendar_time_slots():
    slots = []
    for hour in range(7, 20):
        slot_time = time(hour=hour, minute=0)
        slots.append({
            "value": slot_time.strftime("%H:%M"),
            "label": slot_time.strftime("%I:%M %p").lstrip("0"),
            "time": slot_time,
            "tickets": [],
        })
    return slots


def get_calendar_week_start():
    raw_start = request.args.get("start", "").strip()
    requested = parse_optional_date(raw_start) or date.today()
    return requested - timedelta(days=requested.weekday())


def visible_ticket_or_none(ticket_id, visible_store_numbers):
    try:
        ticket_id = int(ticket_id)
    except (TypeError, ValueError):
        return None

    ticket = MaintenanceTicket.query.get(ticket_id)
    if not ticket:
        return None

    if ticket.store_number not in visible_store_numbers:
        return None

    return ticket


def apply_schedule_form_to_ticket(ticket):
    ticket.store_number = request.form.get("store_number", ticket.store_number).strip()
    ticket.title = request.form.get("title", ticket.title or "").strip()
    ticket.details = request.form.get("details", ticket.details or "").strip()
    ticket.status = request.form.get("status", ticket.status or "open").strip()

    ticket.assigned_to = request.form.get("assigned_to", ticket.assigned_to or "").strip() or None
    ticket.scheduled_date = parse_optional_date(request.form.get("scheduled_date"))
    ticket.scheduled_time = parse_optional_time(request.form.get("scheduled_time"))
    ticket.estimated_minutes = parse_optional_int(request.form.get("estimated_minutes"))
    ticket.priority = normalize_priority(request.form.get("priority"))

    valid_statuses = {"open", "assigned", "in_progress", "complete"}
    if ticket.status not in valid_statuses:
        ticket.status = "open"


@maintenance_bp.route("/", methods=["GET", "POST"])
@login_required
@role_required("admin", "supervisor", "maintenance", "manager")
def index():
    visible_stores = get_visible_stores()
    visible_store_numbers = get_visible_store_numbers()
    role = get_current_role()

    def is_ajax_request():
        return request.headers.get("X-Requested-With") == "XMLHttpRequest"

    def maintenance_response(message, category="success", status_code=200, **extra):
        if is_ajax_request():
            payload = {
                "ok": category != "error",
                "message": message,
                "category": category,
            }
            payload.update(extra)
            return jsonify(payload), status_code

        flash(message, category)
        return redirect(url_for("maintenance.index"))

    if request.method == "POST":
        action = request.form.get("action", "").strip()

        if action == "create":
            store_number = request.form.get("store_number", "").strip()
            title = request.form.get("title", "").strip()
            details = request.form.get("details", "").strip()

            if store_number not in visible_store_numbers:
                return maintenance_response("Invalid store selection.", "error", 400)

            title_lines = split_lines_to_tasks(title)

            if title_lines:
                for line in title_lines:
                    ticket = MaintenanceTicket(
                        store_number=store_number,
                        title=line,
                        details=details,
                        source_type="manual",
                        status="open",
                    )
                    db.session.add(ticket)

                db.session.commit()

                if len(title_lines) == 1:
                    message = "Maintenance task created."
                else:
                    message = f"{len(title_lines)} maintenance tasks created."

                return maintenance_response(message, "success", created_count=len(title_lines))

            if not details:
                return maintenance_response("Task title is required.", "error", 400)

            ticket = MaintenanceTicket(
                store_number=store_number,
                title="General maintenance task",
                details=details,
                source_type="manual",
                status="open",
            )
            db.session.add(ticket)
            db.session.commit()

            return maintenance_response("Maintenance task created.", "success", created_count=1)

        if action == "update":
            ticket_id = request.form.get("ticket_id", "").strip()
            ticket = MaintenanceTicket.query.get(ticket_id)

            if not ticket:
                return maintenance_response("Ticket not found.", "error", 404)

            if ticket.store_number not in visible_store_numbers:
                return maintenance_response("You do not have access to that ticket.", "error", 403)

            store_number = request.form.get("store_number", "").strip()
            title = request.form.get("title", "").strip()
            details = request.form.get("details", "").strip()
            status = request.form.get("status", "").strip()

            valid_statuses = {"open", "assigned", "in_progress", "complete"}

            if store_number not in visible_store_numbers:
                return maintenance_response("Invalid store selection.", "error", 400)

            if not title:
                return maintenance_response("Task title is required.", "error", 400)

            if status not in valid_statuses:
                return maintenance_response("Invalid status.", "error", 400)

            if role == "manager" and store_number != session.get("user_store"):
                return maintenance_response("Managers can only manage tickets for their own store.", "error", 403)

            ticket.store_number = store_number
            ticket.title = title
            ticket.details = details
            ticket.status = status

            db.session.commit()
            return maintenance_response(
                "Maintenance task updated.",
                "success",
                ticket={
                    "id": ticket.id,
                    "store_number": ticket.store_number,
                    "title": ticket.title,
                    "details": ticket.details or "",
                    "status": ticket.status,
                    "status_label": ticket.status.replace("_", " ").title(),
                },
            )

        if action == "delete":
            if role == "manager":
                return maintenance_response("Managers cannot delete maintenance tasks.", "error", 403)

            ticket_id = request.form.get("ticket_id", "").strip()
            ticket = MaintenanceTicket.query.get(ticket_id)

            if not ticket:
                return maintenance_response("Ticket not found.", "error", 404)

            if ticket.store_number not in visible_store_numbers:
                return maintenance_response("You do not have access to that ticket.", "error", 403)

            db.session.delete(ticket)
            db.session.commit()

            return maintenance_response("Maintenance task deleted.", "success", deleted=True, ticket_id=ticket_id)

    status_filter = request.args.get("status", "").strip()
    store_filter = request.args.get("store", "").strip()

    all_visible_tickets = MaintenanceTicket.query.order_by(
        MaintenanceTicket.created_at.asc(),
        MaintenanceTicket.id.asc()
    ).all()

    all_visible_tickets = [t for t in all_visible_tickets if t.store_number in visible_store_numbers]

    status_counts = {
        "open": len([t for t in all_visible_tickets if t.status == "open"]),
        "assigned": len([t for t in all_visible_tickets if t.status == "assigned"]),
        "in_progress": len([t for t in all_visible_tickets if t.status == "in_progress"]),
        "complete": len([t for t in all_visible_tickets if t.status == "complete"]),
    }

    store_counts = {}
    for ticket in all_visible_tickets:
        store_counts[ticket.store_number] = store_counts.get(ticket.store_number, 0) + 1

    tickets = list(all_visible_tickets)

    if status_filter:
        tickets = [t for t in tickets if t.status == status_filter]

    if store_filter:
        tickets = [t for t in tickets if t.store_number == store_filter]

    tickets = sorted(
        tickets,
        key=lambda t: (str(t.store_number or ""), t.created_at or datetime.min, t.id or 0)
    )

    open_tickets = sorted(
        [t for t in tickets if t.status == "open"],
        key=lambda t: (str(t.store_number or ""), t.created_at or datetime.min, t.id or 0)
    )

    assigned_tickets = sorted(
        [t for t in tickets if t.status == "assigned"],
        key=lambda t: (t.created_at or datetime.min, t.id or 0)
    )

    in_progress_tickets = sorted(
        [t for t in tickets if t.status == "in_progress"],
        key=lambda t: (t.created_at or datetime.min, t.id or 0),
        reverse=True
    )

    complete_tickets = sorted(
        [t for t in tickets if t.status == "complete"],
        key=lambda t: (t.created_at or datetime.min, t.id or 0),
        reverse=True
    )

    return render_template(
        "maintenance.html",
        tickets=tickets,
        open_tickets=open_tickets,
        assigned_tickets=assigned_tickets,
        in_progress_tickets=in_progress_tickets,
        complete_tickets=complete_tickets,
        stores=visible_stores,
        status_filter=status_filter,
        store_filter=store_filter,
        status_counts=status_counts,
        store_counts=store_counts,
        all_visible_tickets=all_visible_tickets,
        user_role=role,
    )



@maintenance_bp.route("/calendar", methods=["GET", "POST"])
@login_required
@role_required("admin", "supervisor", "maintenance", "manager")
def calendar():
    visible_stores = get_visible_stores()
    visible_store_numbers = get_visible_store_numbers()
    role = get_current_role()

    maintenance_users = (
        User.query
        .filter(User.role == "maintenance", User.is_active == True)
        .order_by(User.name.asc())
        .all()
    )
    maintenance_names = [u.name.strip() for u in maintenance_users if u.name and u.name.strip()]
    valid_assignees = set(maintenance_names)

    tech_filter = request.args.get("tech", "all").strip() or "all"

    week_start = get_calendar_week_start()
    calendar_start_raw = request.form.get("calendar_start", "").strip()
    if calendar_start_raw:
        posted_week_start = parse_optional_date(calendar_start_raw)
        if posted_week_start:
            week_start = posted_week_start

    week_end = week_start + timedelta(days=6)
    previous_week = week_start - timedelta(days=7)
    next_week = week_start + timedelta(days=7)

    def calendar_redirect():
        args = {"start": week_start.strftime("%Y-%m-%d")}
        if tech_filter and tech_filter != "all":
            args["tech"] = tech_filter
        return redirect(url_for("maintenance.calendar", **args))

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        posted_tech_filter = request.form.get("tech_filter", "").strip()
        if posted_tech_filter:
            tech_filter = posted_tech_filter

        if action == "unschedule":
            if role == "manager":
                flash("Managers can view the maintenance calendar, but cannot schedule maintenance tasks.", "error")
                return calendar_redirect()

            ticket = visible_ticket_or_none(request.form.get("ticket_id"), visible_store_numbers)
            if not ticket:
                flash("Maintenance task not found or not available.", "error")
                return calendar_redirect()

            ticket.scheduled_date = None
            ticket.scheduled_time = None
            ticket.assigned_to = None

            db.session.commit()
            flash("Maintenance task moved back to unscheduled and unassigned.", "success")
            return calendar_redirect()

        if action == "unschedule_all":
            if role == "manager":
                flash("Managers can view the maintenance calendar, but cannot schedule maintenance tasks.", "error")
                return calendar_redirect()

            scheduled_tickets = MaintenanceTicket.query.filter(
                MaintenanceTicket.store_number.in_(visible_store_numbers),
                MaintenanceTicket.scheduled_date.isnot(None),
                MaintenanceTicket.status != "complete"
            ).all()

            for ticket in scheduled_tickets:
                ticket.scheduled_date = None
                ticket.scheduled_time = None
                ticket.assigned_to = None

            db.session.commit()
            flash(f"{len(scheduled_tickets)} maintenance tasks moved back to Unscheduled.", "success")
            return calendar_redirect()

        if action == "schedule":
            if role == "manager":
                flash("Managers can view the maintenance calendar, but cannot schedule maintenance tasks.", "error")
                return calendar_redirect()

            ticket_id_raw = request.form.get("ticket_id", "").strip()
            ticket = None

            if ticket_id_raw:
                ticket = visible_ticket_or_none(ticket_id_raw, visible_store_numbers)
                if not ticket:
                    flash("Maintenance task not found or not available.", "error")
                    return calendar_redirect()

            new_store_number = request.form.get(
                "store_number",
                ticket.store_number if ticket else ""
            ).strip()

            if new_store_number not in visible_store_numbers:
                flash("Invalid store selection.", "error")
                return calendar_redirect()

            title = request.form.get("title", "").strip()
            if not title:
                flash("Task title is required.", "error")
                return calendar_redirect()

            assigned_to = request.form.get("assigned_to", "").strip()
            if assigned_to and assigned_to not in valid_assignees:
                flash("Invalid maintenance assignment.", "error")
                return calendar_redirect()

            is_new_ticket = ticket is None

            if is_new_ticket:
                ticket = MaintenanceTicket(
                    store_number=new_store_number,
                    title=title,
                    status="open",
                    priority="normal",
                )
                db.session.add(ticket)

            apply_schedule_form_to_ticket(ticket)

            if ticket.scheduled_time and not ticket.scheduled_date:
                flash("Please choose a scheduled date when setting a scheduled time.", "error")
                return calendar_redirect()

            if ticket.assigned_to and ticket.status == "open":
                ticket.status = "assigned"

            db.session.commit()

            if is_new_ticket:
                flash("Maintenance task created.", "success")
            else:
                flash("Maintenance schedule updated.", "success")

            return calendar_redirect()

    all_tickets = MaintenanceTicket.query.order_by(
        MaintenanceTicket.created_at.asc(),
        MaintenanceTicket.id.asc()
    ).all()

    all_tickets = [
        t for t in all_tickets
        if t.store_number in visible_store_numbers
    ]

    def ticket_matches_calendar_filter(ticket):
        assigned_to = (ticket.assigned_to or "").strip()

        if tech_filter == "all":
            return True

        if tech_filter == "unassigned":
            return not assigned_to

        return assigned_to == tech_filter

    tickets = [t for t in all_tickets if ticket_matches_calendar_filter(t)]

    def ticket_matches_unscheduled_dispatch(ticket):
        assigned_to = (ticket.assigned_to or "").strip()

        if tech_filter == "all":
            return True

        if tech_filter == "unassigned":
            return not assigned_to

        # When viewing Nick/Jim, keep unassigned tasks visible so they can be dragged
        # directly onto that tech's calendar and assigned in one move.
        return (not assigned_to) or assigned_to == tech_filter

    def ticket_sort_key(ticket):
        return (
            ticket.scheduled_date or date.max,
            ticket.scheduled_time or time(hour=23, minute=59),
            ticket.created_at or datetime.min,
            ticket.id or 0,
        )

    tickets = sorted(tickets, key=ticket_sort_key)

    days = []
    for offset in range(7):
        current_day = week_start + timedelta(days=offset)
        day_tickets = [t for t in tickets if t.scheduled_date == current_day]
        any_time_tickets = [t for t in day_tickets if not t.scheduled_time]

        slots = build_calendar_time_slots()
        for slot in slots:
            slot["tickets"] = [
                t for t in day_tickets
                if t.scheduled_time and t.scheduled_time.hour == slot["time"].hour
            ]

        days.append({
            "date": current_day,
            "tickets": day_tickets,
            "any_time_tickets": any_time_tickets,
            "slots": slots,
        })

    unscheduled_tickets = [
        t for t in all_tickets
        if (
            not t.scheduled_date
            and t.status != "complete"
            and ticket_matches_unscheduled_dispatch(t)
        )
    ]

    def unscheduled_sort_key(ticket):
        try:
            store_sort = int(ticket.store_number)
        except (TypeError, ValueError):
            store_sort = 999999

        status_order = {
            "open": 1,
            "assigned": 2,
            "in_progress": 3,
            "complete": 4,
        }

        return (
            store_sort,
            status_order.get(ticket.status, 99),
            ticket.created_at or datetime.min,
            ticket.id or 0,
        )

    unscheduled_tickets = sorted(unscheduled_tickets, key=unscheduled_sort_key)

    return render_template(
        "maintenance_calendar.html",
        stores=visible_stores,
        days=days,
        unscheduled_tickets=unscheduled_tickets,
        maintenance_users=maintenance_users,
        tech_filter=tech_filter,
        week_start=week_start,
        week_end=week_end,
        previous_week=previous_week,
        next_week=next_week,
        user_role=role,
        format_time=format_maintenance_time,
    )


@maintenance_bp.route("/calendar/move", methods=["POST"])
@login_required
@role_required("admin", "supervisor", "maintenance")
def move_calendar_ticket():
    visible_store_numbers = get_visible_store_numbers()

    ticket = visible_ticket_or_none(request.form.get("ticket_id"), visible_store_numbers)
    if not ticket:
        return jsonify({"ok": False, "message": "Task not found or access denied."}), 404

    maintenance_users = (
        User.query
        .filter(User.role == "maintenance", User.is_active == True)
        .order_by(User.name.asc())
        .all()
    )
    valid_assignees = {u.name.strip() for u in maintenance_users if u.name and u.name.strip()}

    scheduled_date_raw = request.form.get("scheduled_date", "").strip()
    scheduled_time_raw = request.form.get("scheduled_time", "").strip()
    assigned_to_raw = request.form.get("assigned_to", "").strip()
    should_update_assignment = request.form.get("update_assignment", "").strip() == "1"

    if assigned_to_raw and assigned_to_raw not in valid_assignees:
        return jsonify({"ok": False, "message": "Invalid maintenance assignment."}), 400

    if scheduled_date_raw == "unscheduled":
        ticket.scheduled_date = None
        ticket.scheduled_time = None
        db.session.commit()
        return jsonify({"ok": True, "message": "Task moved to unscheduled."})

    scheduled_date = parse_optional_date(scheduled_date_raw)
    if not scheduled_date:
        return jsonify({"ok": False, "message": "Invalid scheduled date."}), 400

    ticket.scheduled_date = scheduled_date
    ticket.scheduled_time = parse_optional_time(scheduled_time_raw)

    # When viewing a specific tech filter, dragging onto the calendar assigns it to that tech.
    # When viewing All, drag only changes date/time and preserves the current assignment.
    if should_update_assignment:
        ticket.assigned_to = assigned_to_raw or None

    if ticket.assigned_to and ticket.status == "open":
        ticket.status = "assigned"

    db.session.commit()
    return jsonify({"ok": True, "message": "Maintenance task moved."})


@maintenance_bp.route("/time-card", methods=["GET", "POST"])
@login_required
@role_required("maintenance", "admin", "hr")
def time_card():
    role = get_current_role()

    if role in ["admin", "supervisor"]:
        return redirect(url_for("maintenance.time_cards"))

    user = get_current_user_record()
    if not user or user.role != "maintenance":
        flash("Maintenance time card is only available for maintenance users.", "error")
        return redirect(url_for("dashboard.home"))

    work_date = today_et_local()
    card = get_or_create_time_card(user.id, work_date)

    if request.method == "POST":
        action = request.form.get("action", "").strip()

        if action == "clock_in":
            if card.clock_in_at and not card.clock_out_at:
                flash("You are already clocked in.", "error")
            elif card.clock_in_at and card.clock_out_at:
                flash("Today's time card is already completed. Contact admin if it needs an edit.", "error")
            else:
                card.clock_in_at = now_utc_naive()
                flash("Clocked in.", "success")

        elif action == "clock_out":
            if not card.clock_in_at:
                flash("You need to clock in before clocking out.", "error")
            elif card.clock_out_at:
                flash("You are already clocked out.", "error")
            else:
                card.clock_out_at = now_utc_naive()
                flash("Clocked out.", "success")

        elif action == "save_note":
            card.notes = request.form.get("notes", "").strip()
            flash("Time card note saved.", "success")

        db.session.commit()
        return redirect(url_for("maintenance.time_card"))

    today_hours = hours_between(card.clock_in_at, card.clock_out_at)
    is_clocked_in = bool(card.clock_in_at and not card.clock_out_at)

    recent_cards = (
        MaintenanceTimeCard.query
        .filter(MaintenanceTimeCard.user_id == user.id)
        .order_by(MaintenanceTimeCard.work_date.desc())
        .limit(14)
        .all()
    )

    return render_template(
        "maintenance_time_card.html",
        card=card,
        user=user,
        work_date=work_date,
        today_hours=today_hours,
        is_clocked_in=is_clocked_in,
        recent_cards=recent_cards,
        format_dt_et=format_dt_et,
        hours_between=hours_between,
    )


@maintenance_bp.route("/time-cards", methods=["GET", "POST"])
@login_required
@role_required("admin", "maintenance", "hr")
def time_cards():
    role = get_current_role()
    current_user = get_current_user_record()

    today = today_et_local()
    default_start = monday_for_date(today)
    default_end = sunday_for_date(today)

    start_date = parse_optional_date(request.args.get("start")) or default_start
    end_date = parse_optional_date(request.args.get("end")) or default_end

    if end_date < start_date:
        end_date = start_date

    tech_raw = request.args.get("tech", "all").strip()
    tech_id = None

    if role == "maintenance":
        if not current_user or current_user.role != "maintenance":
            flash("Maintenance time cards are only available for maintenance users.", "error")
            return redirect(url_for("dashboard.home"))
        tech_id = current_user.id
        tech_raw = str(current_user.id)
    elif tech_raw and tech_raw != "all":
        try:
            tech_id = int(tech_raw)
        except ValueError:
            tech_id = None

    if request.method == "POST":
        action = request.form.get("action", "").strip()

        if action == "save_card":
            card_id_raw = request.form.get("card_id", "").strip()
            user_id_raw = request.form.get("user_id", "").strip()
            work_date = parse_optional_date(request.form.get("work_date"))

            if role == "maintenance":
                user = current_user
            else:
                try:
                    user_id = int(user_id_raw)
                except (TypeError, ValueError):
                    user_id = None
                user = User.query.get(user_id) if user_id else None

            if not user or user.role != "maintenance":
                flash("Invalid maintenance user.", "error")
                return redirect(url_for("maintenance.time_cards", start=start_date, end=end_date, tech=tech_raw or "all"))

            if not work_date:
                flash("Work date is required.", "error")
                return redirect(url_for("maintenance.time_cards", start=start_date, end=end_date, tech=tech_raw or "all"))

            if card_id_raw:
                card = MaintenanceTimeCard.query.get(card_id_raw)
                if not card:
                    flash("Time card not found.", "error")
                    return redirect(url_for("maintenance.time_cards", start=start_date, end=end_date, tech=tech_raw or "all"))

                if role == "maintenance" and card.user_id != current_user.id:
                    flash("You can only edit your own time cards.", "error")
                    return redirect(url_for("maintenance.time_cards", start=start_date, end=end_date, tech=tech_raw or "all"))
            else:
                card = get_or_create_time_card(user.id, work_date)

            clock_in_at = parse_datetime_local(request.form.get("clock_in_at"))
            clock_out_at = parse_datetime_local(request.form.get("clock_out_at"))

            if clock_in_at and clock_out_at and clock_out_at <= clock_in_at:
                flash("Clock out cannot be before clock in.", "error")
                return redirect(url_for("maintenance.time_cards", start=start_date, end=end_date, tech=tech_raw or "all"))

            card.user_id = user.id
            card.work_date = work_date
            card.clock_in_at = clock_in_at
            card.clock_out_at = clock_out_at
            card.notes = request.form.get("notes", "").strip()

            # Any save from the report/edit modal is a manual adjustment.
            card.is_edited = True
            card.edited_at = now_utc_naive()
            if current_user:
                card.edited_by_user_id = current_user.id

            db.session.commit()
            flash("Time card saved and marked as edited.", "success")
            return redirect(url_for("maintenance.time_cards", start=start_date, end=end_date, tech=tech_raw or "all"))

    report = build_time_card_report(start_date, end_date, tech_id=tech_id)
    maintenance_users = get_maintenance_users()

    this_week_start = monday_for_date(today)
    this_week_end = sunday_for_date(today)
    last_week_start = this_week_start - timedelta(days=7)
    last_week_end = this_week_start - timedelta(days=1)
    last_14_start = today - timedelta(days=13)
    last_14_end = today

    return render_template(
        "maintenance_time_cards.html",
        start_date=start_date,
        end_date=end_date,
        tech_raw=tech_raw,
        tech_id=tech_id,
        maintenance_users=maintenance_users,
        summaries=report["summaries"],
        rows=report["rows"],
        total_summary=report["total_summary"],
        format_dt_et=format_dt_et,
        today=today,
        user_role=role,
        current_user=current_user,
        this_week_start=this_week_start,
        this_week_end=this_week_end,
        last_week_start=last_week_start,
        last_week_end=last_week_end,
        last_14_start=last_14_start,
        last_14_end=last_14_end,
    )


@maintenance_bp.route("/time-cards/pdf")
@login_required
@role_required("admin", "maintenance", "hr")
def time_cards_pdf():
    role = get_current_role()
    current_user = get_current_user_record()

    today = today_et_local()
    default_start = monday_for_date(today)
    default_end = sunday_for_date(today)

    start_date = parse_optional_date(request.args.get("start")) or default_start
    end_date = parse_optional_date(request.args.get("end")) or default_end

    if end_date < start_date:
        end_date = start_date

    tech_raw = request.args.get("tech", "all").strip()
    tech_id = None

    if role == "maintenance":
        if not current_user or current_user.role != "maintenance":
            flash("Maintenance time card export is only available for maintenance users.", "error")
            return redirect(url_for("dashboard.home"))
        tech_id = current_user.id
    elif tech_raw and tech_raw != "all":
        try:
            tech_id = int(tech_raw)
        except ValueError:
            tech_id = None

    pdf_buffer = create_time_cards_pdf(start_date, end_date, tech_id=tech_id)

    filename = f"maintenance_time_cards_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


@maintenance_bp.route("/export/excel")
@login_required
@role_required("admin", "supervisor", "maintenance", "manager")
def export_excel():
    tickets, _, status_filter, store_filter = get_filtered_tickets()
    workbook_stream = create_maintenance_excel(tickets, status_filter, store_filter)

    filename_parts = ["maintenance_export"]
    if store_filter:
        filename_parts.append(store_filter)
    if status_filter:
        filename_parts.append(status_filter)

    filename = "_".join(filename_parts) + ".xlsx"

    return send_file(
        workbook_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )