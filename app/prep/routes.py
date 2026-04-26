from collections import defaultdict
from datetime import datetime
import json
import os
import re
from uuid import uuid4

from flask import Blueprint, current_app, render_template, request, redirect, url_for, flash, session, jsonify
from openpyxl import load_workbook

from app.auth.routes import login_required, role_required
from app.extensions import db
from app.models import PrepTemplateItem, DailyPrep, DailyPrepItem, Store, today_et

prep_bp = Blueprint("prep", __name__, url_prefix="/prep")


SECTION_OPTIONS = [
    "Sauces",
    "Cheese",
    "Chicken / Wings / Boneless",
    "Veggie / Specialty / Other",
]


DEFAULT_IMPORT_SECTION = "Veggie / Specialty / Other"


def weekday_field_name(prep_date):
    return prep_date.strftime("%A").lower()


def get_visible_stores():
    role = session.get("user_role")
    user_area = session.get("user_area")
    user_store = session.get("user_store")

    if role == "admin":
        return Store.query.filter_by(is_active=True).order_by(Store.store_number.asc()).all()

    if role == "supervisor":
        return Store.query.filter_by(
            area_name=user_area,
            is_active=True
        ).order_by(Store.store_number.asc()).all()

    if role == "manager":
        return Store.query.filter_by(
            store_number=user_store,
            is_active=True
        ).order_by(Store.store_number.asc()).all()

    return []


def get_allowed_store_numbers():
    return {store.store_number for store in get_visible_stores()}


def normalize_item_key(value):
    return (value or "").strip().lower()


def normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def clean_optional_text(value):
    cleaned = (value or "").strip()
    return cleaned or None


def safe_float(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    cleaned = str(value).strip()
    if not cleaned:
        return None

    cleaned = cleaned.replace(",", "").replace("$", "")
    cleaned = re.sub(r"[^0-9.\-]", "", cleaned)

    if cleaned in {"", "-", ".", "-."}:
        return None

    try:
        return float(cleaned)
    except ValueError:
        return None


def format_usage_value(value, unit=None):
    if value is None:
        return ""

    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)

    if number.is_integer():
        formatted = str(int(number))
    else:
        formatted = f"{number:.2f}".rstrip("0").rstrip(".")

    unit_text = (unit or "").strip()
    if unit_text:
        return f"{formatted} {unit_text}"

    return formatted


def split_store_item(raw_item, raw_store=None):
    item_text = str(raw_item or "").strip()
    store_text = str(raw_store or "").strip()

    if store_text:
        store_match = re.search(r"\d{4}", store_text)
        if store_match:
            return store_match.group(0), item_text

    combined_match = re.match(r"^\s*(\d{4})\s*[-–—]\s*(.+?)\s*$", item_text)
    if combined_match:
        return combined_match.group(1), combined_match.group(2).strip()

    return None, item_text


def find_header_indexes(header_values):
    normalized = [normalize_header(value) for value in header_values]

    ideal_index = None
    item_index = None
    store_index = None
    unit_index = None
    actual_index = None

    for index, header in enumerate(normalized):
        if not header:
            continue

        if ideal_index is None and header in {"idealusage", "idealqty", "idealquantity"}:
            ideal_index = index

        if item_index is None and header in {
            "item",
            "itemname",
            "inventoryitem",
            "inventoryitemname",
            "description",
            "product",
            "productname",
        }:
            item_index = index

        if store_index is None and header in {
            "store",
            "storenumber",
            "store#",
            "location",
            "locationnumber",
        }:
            store_index = index

        if unit_index is None and header in {
            "unit",
            "units",
            "uom",
            "usageunit",
        }:
            unit_index = index

        if actual_index is None and header in {
            "actualusage",
            "actualqty",
            "actualquantity",
            "usage",
        }:
            actual_index = index

    if ideal_index is None:
        for index, header in enumerate(normalized):
            if "ideal" in header and "usage" in header:
                ideal_index = index
                break

    if item_index is None:
        for index, header in enumerate(normalized):
            if "item" in header or "description" in header or "product" in header:
                item_index = index
                break

    if store_index is None:
        for index, header in enumerate(normalized):
            if "store" in header or "location" in header:
                store_index = index
                break

    if unit_index is None:
        for index, header in enumerate(normalized):
            if "unit" in header or "uom" in header:
                unit_index = index
                break

    if actual_index is None:
        for index, header in enumerate(normalized):
            if "actual" in header and "usage" in header:
                actual_index = index
                break

    return {
        "ideal_index": ideal_index,
        "item_index": item_index,
        "store_index": store_index,
        "unit_index": unit_index,
        "actual_index": actual_index,
    }


def build_template_lookup(allowed_store_numbers):
    existing_template_items = PrepTemplateItem.query.filter(
        PrepTemplateItem.store_number.in_(allowed_store_numbers)
    ).all()

    template_lookup = {}

    for template in existing_template_items:
        keys = {
            normalize_item_key(template.item_name),
            normalize_item_key(template.report_item_name),
        }

        for key in keys:
            if key:
                template_lookup[(template.store_number, key)] = template

    return template_lookup


def refresh_preview_match_status(rows, allowed_store_numbers):
    template_lookup = build_template_lookup(allowed_store_numbers)
    matched_count = 0
    refreshed_rows = []

    for row in rows:
        store_number = row.get("store_number")
        report_item_name = row.get("report_item_name", "")
        item_key = normalize_item_key(report_item_name)

        matched_template = template_lookup.get((store_number, item_key))
        status = "New"

        if matched_template:
            matched_count += 1
            status = "Matched"

        refreshed = dict(row)
        refreshed["matched_item_name"] = matched_template.item_name if matched_template else ""
        refreshed["section_name"] = matched_template.section_name if matched_template else ""
        refreshed["status"] = status
        refreshed_rows.append(refreshed)

    return refreshed_rows, matched_count


def parse_ideal_usage_workbook(file_storage, allowed_store_numbers):
    workbook = load_workbook(file_storage, data_only=True, read_only=True)
    sheet = workbook.active

    header_row_number = None
    indexes = None

    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        row_values = list(row)
        possible_indexes = find_header_indexes(row_values)

        if possible_indexes["ideal_index"] is not None:
            header_row_number = row_number
            indexes = possible_indexes
            break

    if header_row_number is None or indexes is None:
        raise ValueError("Could not find an Ideal Usage column in the uploaded file.")

    ideal_index = indexes["ideal_index"]
    item_index = indexes["item_index"]
    store_index = indexes["store_index"]
    unit_index = indexes["unit_index"]
    actual_index = indexes["actual_index"]

    if item_index is None:
        item_index = 0

    preview_rows = []
    skipped_rows = 0

    template_lookup = build_template_lookup(allowed_store_numbers)
    matched_template_count = 0

    for row in sheet.iter_rows(min_row=header_row_number + 1, values_only=True):
        row_values = list(row)

        raw_item = row_values[item_index] if item_index is not None and item_index < len(row_values) else None
        raw_store = row_values[store_index] if store_index is not None and store_index < len(row_values) else None
        raw_ideal = row_values[ideal_index] if ideal_index is not None and ideal_index < len(row_values) else None
        raw_unit = row_values[unit_index] if unit_index is not None and unit_index < len(row_values) else None
        raw_actual = row_values[actual_index] if actual_index is not None and actual_index < len(row_values) else None

        if raw_item is None and raw_ideal is None:
            continue

        ideal_usage = safe_float(raw_ideal)
        if ideal_usage is None:
            skipped_rows += 1
            continue

        store_number, item_name = split_store_item(raw_item, raw_store)

        if not store_number or not item_name:
            skipped_rows += 1
            continue

        if store_number not in allowed_store_numbers:
            skipped_rows += 1
            continue

        item_key = normalize_item_key(item_name)
        matched_template = template_lookup.get((store_number, item_key))
        matched_status = "New"

        if matched_template:
            matched_template_count += 1
            matched_status = "Matched"

        preview_rows.append({
            "store_number": store_number,
            "report_item_name": item_name,
            "matched_item_name": matched_template.item_name if matched_template else "",
            "section_name": matched_template.section_name if matched_template else "",
            "unit": str(raw_unit or "").strip(),
            "actual_usage": safe_float(raw_actual),
            "ideal_usage": ideal_usage,
            "status": matched_status,
        })

    preview_rows.sort(key=lambda row: (row["store_number"], row["report_item_name"].lower()))

    return {
        "rows": preview_rows,
        "row_count": len(preview_rows),
        "skipped_count": skipped_rows,
        "matched_count": matched_template_count,
        "new_count": max(len(preview_rows) - matched_template_count, 0),
        "sheet_name": sheet.title,
    }


def get_preview_storage_dir():
    storage_dir = os.path.join(current_app.instance_path, "prep_import_previews")
    os.makedirs(storage_dir, exist_ok=True)
    return storage_dir


def save_upload_preview(upload_preview):
    token = uuid4().hex
    storage_path = os.path.join(get_preview_storage_dir(), f"{token}.json")

    payload = {
        "created_at": datetime.utcnow().isoformat(),
        "sheet_name": upload_preview.get("sheet_name"),
        "rows": upload_preview.get("rows", []),
        "skipped_count": upload_preview.get("skipped_count", 0),
    }

    with open(storage_path, "w", encoding="utf-8") as file:
        json.dump(payload, file)

    return token


def load_upload_preview(token):
    if not token or not re.fullmatch(r"[a-f0-9]{32}", token):
        return None

    storage_path = os.path.join(get_preview_storage_dir(), f"{token}.json")

    if not os.path.exists(storage_path):
        return None

    with open(storage_path, "r", encoding="utf-8") as file:
        return json.load(file)


def build_upload_preview_from_saved(saved_preview, allowed_store_numbers, token=None):
    rows = saved_preview.get("rows", [])
    refreshed_rows, matched_count = refresh_preview_match_status(rows, allowed_store_numbers)

    return {
        "token": token,
        "rows": refreshed_rows,
        "row_count": len(refreshed_rows),
        "skipped_count": saved_preview.get("skipped_count", 0),
        "matched_count": matched_count,
        "new_count": max(len(refreshed_rows) - matched_count, 0),
        "sheet_name": saved_preview.get("sheet_name", "Uploaded File"),
    }


def import_selected_preview_rows(saved_preview, selected_indexes, import_mode, allowed_store_numbers):
    rows = saved_preview.get("rows", [])

    if import_mode not in {"create_missing", "update_matched", "create_and_update"}:
        import_mode = "create_missing"

    selected_index_set = set()

    for raw_index in selected_indexes:
        try:
            selected_index_set.add(int(raw_index))
        except (TypeError, ValueError):
            continue

    if not selected_index_set:
        return {
            "created_count": 0,
            "updated_count": 0,
            "skipped_count": 0,
            "selected_count": 0,
        }

    template_lookup = build_template_lookup(allowed_store_numbers)

    created_count = 0
    updated_count = 0
    skipped_count = 0
    selected_count = 0

    for row_index, row in enumerate(rows):
        if row_index not in selected_index_set:
            continue

        selected_count += 1

        store_number = row.get("store_number")
        report_item_name = (row.get("report_item_name") or "").strip()
        unit = (row.get("unit") or "").strip()
        ideal_usage = row.get("ideal_usage")

        if not store_number or store_number not in allowed_store_numbers or not report_item_name:
            skipped_count += 1
            continue

        item_key = normalize_item_key(report_item_name)
        existing_item = template_lookup.get((store_number, item_key))

        build_to_value = format_usage_value(ideal_usage, unit)

        if existing_item:
            if import_mode in {"update_matched", "create_and_update"}:
                existing_item.report_item_name = report_item_name
                existing_item.prep_unit = unit or existing_item.prep_unit
                existing_item.build_to = build_to_value or existing_item.build_to
                updated_count += 1
            else:
                skipped_count += 1
            continue

        if import_mode in {"create_missing", "create_and_update"}:
            new_item = PrepTemplateItem(
                store_number=store_number,
                section_name=DEFAULT_IMPORT_SECTION,
                item_name=report_item_name,
                build_to=build_to_value or None,
                instructions=None,
                report_item_name=report_item_name,
                prep_unit=unit or None,
                rounding_increment=None,
                minimum_build=None,
                buffer_percent=None,
                conversion_notes=None,
                monday=True,
                tuesday=True,
                wednesday=True,
                thursday=True,
                friday=True,
                saturday=True,
                sunday=True,
                sort_order=999,
                is_active=True,
            )
            db.session.add(new_item)
            created_count += 1
        else:
            skipped_count += 1

    db.session.commit()

    return {
        "created_count": created_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "selected_count": selected_count,
    }


def sync_missing_daily_prep_items(daily):
    weekday_name = weekday_field_name(daily.prep_date)

    template_items = PrepTemplateItem.query.filter_by(
        store_number=daily.store_number,
        is_active=True
    ).order_by(
        PrepTemplateItem.section_name.asc(),
        PrepTemplateItem.sort_order.asc(),
        PrepTemplateItem.id.asc()
    ).all()

    existing_template_ids = {
        item.template_item_id
        for item in daily.items
        if item.template_item_id is not None
    }

    added_any = False

    for template in template_items:
        if not getattr(template, weekday_name, False):
            continue

        if template.id in existing_template_ids:
            continue

        db.session.add(
            DailyPrepItem(
                daily_prep_id=daily.id,
                template_item_id=template.id,
                section_name=template.section_name,
                item_name=template.item_name,
                build_to=template.build_to,
                instructions=template.instructions,
                is_completed=False,
                completed_at=None,
            )
        )
        added_any = True

    if added_any:
        db.session.commit()

    return added_any


def get_or_create_daily_prep(store_number, prep_date):
    daily = DailyPrep.query.filter_by(
        store_number=store_number,
        prep_date=prep_date
    ).first()

    if not daily:
        daily = DailyPrep(
            store_number=store_number,
            prep_date=prep_date,
        )
        db.session.add(daily)
        db.session.flush()

        weekday_name = weekday_field_name(prep_date)

        template_items = PrepTemplateItem.query.filter_by(
            store_number=store_number,
            is_active=True
        ).order_by(
            PrepTemplateItem.section_name.asc(),
            PrepTemplateItem.sort_order.asc(),
            PrepTemplateItem.id.asc()
        ).all()

        for template in template_items:
            if not getattr(template, weekday_name, False):
                continue

            db.session.add(
                DailyPrepItem(
                    daily_prep_id=daily.id,
                    template_item_id=template.id,
                    section_name=template.section_name,
                    item_name=template.item_name,
                    build_to=template.build_to,
                    instructions=template.instructions,
                    is_completed=False,
                    completed_at=None,
                )
            )

        db.session.commit()
        return daily

    sync_missing_daily_prep_items(daily)
    return daily


def build_grouped_items(items):
    grouped = defaultdict(list)

    for item in items:
        grouped[item.section_name].append(item)

    ordered = {}
    for section in SECTION_OPTIONS:
        ordered[section] = grouped.get(section, [])

    for section_name in sorted(grouped.keys()):
        if section_name not in ordered:
            ordered[section_name] = grouped[section_name]

    return ordered


@prep_bp.route("/", methods=["GET"])
@login_required
@role_required("admin", "supervisor", "manager")
def index():
    visible_stores = get_visible_stores()

    if not visible_stores:
        flash("No stores are assigned to this user.", "error")
        return render_template(
            "placeholder.html",
            page_title="Prep Module",
            page_message="No stores are assigned to this user."
        )

    default_store = visible_stores[0].store_number
    requested_store = request.args.get("store", default_store).strip()
    allowed_store_numbers = {store.store_number for store in visible_stores}

    store_number = requested_store if requested_store in allowed_store_numbers else default_store

    requested_date_str = request.args.get("date", "").strip()
    today = today_et()

    if requested_date_str:
        try:
            selected_date = datetime.strptime(requested_date_str, "%Y-%m-%d").date()
        except ValueError:
            selected_date = today
    else:
        selected_date = today

    is_read_only = selected_date < today

    daily = get_or_create_daily_prep(store_number, selected_date)
    grouped_items = build_grouped_items(
        sorted(daily.items, key=lambda x: (x.section_name, x.item_name, x.id))
    )

    total_items = len(daily.items)
    completed_items = sum(1 for item in daily.items if item.is_completed)

    history = DailyPrep.query.filter_by(
        store_number=store_number
    ).order_by(DailyPrep.prep_date.desc()).limit(14).all()

    return render_template(
        "prep/index.html",
        daily=daily,
        grouped_items=grouped_items,
        store_number=store_number,
        selected_date=selected_date.strftime("%Y-%m-%d"),
        today_label=selected_date.strftime("%B %d, %Y"),
        stores=visible_stores,
        history=history,
        total_items=total_items,
        completed_items=completed_items,
        is_read_only=is_read_only,
    )


@prep_bp.route("/autosave-item", methods=["POST"])
@login_required
@role_required("admin", "supervisor", "manager")
def autosave_item():
    data = request.get_json() or {}

    item_id = data.get("item_id")
    is_completed = bool(data.get("is_completed", False))

    item = DailyPrepItem.query.get(item_id)
    if not item:
        return jsonify({"success": False, "error": "Prep item not found."}), 404

    allowed_store_numbers = get_allowed_store_numbers()
    if item.daily_prep.store_number not in allowed_store_numbers:
        return jsonify({"success": False, "error": "Unauthorized."}), 403

    if item.daily_prep.prep_date < today_et():
        return jsonify({"success": False, "error": "Past prep sheets are read-only."}), 400

    item.is_completed = is_completed
    item.completed_at = datetime.utcnow() if is_completed else None
    db.session.commit()

    daily = item.daily_prep
    total_items = len(daily.items)
    completed_items = sum(1 for row in daily.items if row.is_completed)

    return jsonify({
        "success": True,
        "completed_items": completed_items,
        "total_items": total_items,
    })


@prep_bp.route("/manage-autosave", methods=["POST"])
@login_required
@role_required("admin", "supervisor")
def manage_autosave():
    data = request.get_json() or {}

    item_id = data.get("item_id")
    store_number = (data.get("store_number") or "").strip()

    if not item_id or not store_number:
        return jsonify({"success": False, "error": "Missing item/store."}), 400

    allowed_store_numbers = get_allowed_store_numbers()
    if store_number not in allowed_store_numbers:
        return jsonify({"success": False, "error": "Unauthorized."}), 403

    item = PrepTemplateItem.query.get(item_id)
    if not item or item.store_number != store_number:
        return jsonify({"success": False, "error": "Prep item not found."}), 404

    section_name = (data.get("section_name") or "").strip()
    item_name = (data.get("item_name") or "").strip()
    build_to = (data.get("build_to") or "").strip()
    instructions = (data.get("instructions") or "").strip()

    if not section_name or not item_name:
        return jsonify({"success": False, "error": "Section and item name are required."}), 400

    try:
        sort_order = int(str(data.get("sort_order", "0")).strip() or "0")
    except ValueError:
        return jsonify({"success": False, "error": "Sort order must be a number."}), 400

    item.section_name = section_name
    item.item_name = item_name
    item.build_to = build_to or None
    item.instructions = instructions or None
    item.sort_order = sort_order

    item.report_item_name = clean_optional_text(data.get("report_item_name"))
    item.prep_unit = clean_optional_text(data.get("prep_unit"))
    item.rounding_increment = clean_optional_text(data.get("rounding_increment"))
    item.minimum_build = clean_optional_text(data.get("minimum_build"))
    item.buffer_percent = clean_optional_text(data.get("buffer_percent"))
    item.conversion_notes = clean_optional_text(data.get("conversion_notes"))

    item.monday = bool(data.get("monday", False))
    item.tuesday = bool(data.get("tuesday", False))
    item.wednesday = bool(data.get("wednesday", False))
    item.thursday = bool(data.get("thursday", False))
    item.friday = bool(data.get("friday", False))
    item.saturday = bool(data.get("saturday", False))
    item.sunday = bool(data.get("sunday", False))
    item.is_active = bool(data.get("is_active", False))

    db.session.commit()

    return jsonify({"success": True})


@prep_bp.route("/manage", methods=["GET", "POST"])
@login_required
@role_required("admin", "supervisor")
def manage():
    visible_stores = get_visible_stores()

    if not visible_stores:
        flash("No stores are assigned to this user.", "error")
        return render_template(
            "placeholder.html",
            page_title="Prep Admin",
            page_message="No stores are assigned to this user."
        )

    default_store = visible_stores[0].store_number
    requested_store = request.args.get("store", default_store).strip()
    allowed_store_numbers = {store.store_number for store in visible_stores}

    store_number = requested_store if requested_store in allowed_store_numbers else default_store
    show_inactive = request.args.get("show_inactive", "").strip().lower() in {"1", "true", "yes", "on"}

    upload_preview = None

    if request.method == "POST":
        action = request.form.get("action", "").strip()

        if action == "preview_ideal_usage":
            upload_file = request.files.get("ideal_usage_file")

            if not upload_file or not upload_file.filename:
                flash("Choose an Excel file to preview.", "error")
                return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

            filename = upload_file.filename.lower()
            if not filename.endswith((".xlsx", ".xlsm")):
                flash("Upload an .xlsx or .xlsm Excel file.", "error")
                return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

            try:
                upload_preview = parse_ideal_usage_workbook(upload_file, allowed_store_numbers)
                preview_token = save_upload_preview(upload_preview)
                upload_preview["token"] = preview_token

                flash(
                    f"Preview loaded: {upload_preview['row_count']} rows found, "
                    f"{upload_preview['matched_count']} matched to existing prep items.",
                    "success"
                )
            except Exception as error:
                flash(f"Could not preview ideal usage file: {error}", "error")
                upload_preview = None

        elif action == "import_ideal_usage_selected":
            preview_token = request.form.get("preview_token", "").strip()
            import_mode = request.form.get("import_mode", "create_missing").strip()
            selected_rows = request.form.getlist("selected_rows")

            saved_preview = load_upload_preview(preview_token)

            if not saved_preview:
                flash("Upload preview expired or could not be found. Please upload the file again.", "error")
                return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

            if not selected_rows:
                upload_preview = build_upload_preview_from_saved(saved_preview, allowed_store_numbers, preview_token)
                flash("Choose at least one row to import.", "error")
            else:
                result = import_selected_preview_rows(
                    saved_preview=saved_preview,
                    selected_indexes=selected_rows,
                    import_mode=import_mode,
                    allowed_store_numbers=allowed_store_numbers,
                )

                flash(
                    f"Import complete: {result['created_count']} created, "
                    f"{result['updated_count']} updated, {result['skipped_count']} skipped.",
                    "success"
                )

                return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

        elif action == "create":
            section_name = request.form.get("section_name", "").strip()
            item_name = request.form.get("item_name", "").strip()
            build_to = request.form.get("build_to", "").strip()
            instructions = request.form.get("instructions", "").strip()
            sort_order_raw = request.form.get("sort_order", "0").strip()

            if not section_name or not item_name:
                flash("Section and item name are required.", "error")
                return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

            try:
                sort_order = int(sort_order_raw or "0")
            except ValueError:
                flash("Sort order must be a number.", "error")
                return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

            item = PrepTemplateItem(
                store_number=store_number,
                section_name=section_name,
                item_name=item_name,
                build_to=build_to or None,
                instructions=instructions or None,
                report_item_name=clean_optional_text(request.form.get("report_item_name")),
                prep_unit=clean_optional_text(request.form.get("prep_unit")),
                rounding_increment=clean_optional_text(request.form.get("rounding_increment")),
                minimum_build=clean_optional_text(request.form.get("minimum_build")),
                buffer_percent=clean_optional_text(request.form.get("buffer_percent")),
                conversion_notes=clean_optional_text(request.form.get("conversion_notes")),
                monday=request.form.get("monday") == "on",
                tuesday=request.form.get("tuesday") == "on",
                wednesday=request.form.get("wednesday") == "on",
                thursday=request.form.get("thursday") == "on",
                friday=request.form.get("friday") == "on",
                saturday=request.form.get("saturday") == "on",
                sunday=request.form.get("sunday") == "on",
                sort_order=sort_order,
                is_active=request.form.get("is_active") == "on",
            )
            db.session.add(item)
            db.session.commit()
            flash("Prep item created.", "success")
            return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

        elif action == "copy_store":
            source_store = request.form.get("source_store", "").strip()
            destination_store = request.form.get("destination_store", "").strip()
            copy_mode = request.form.get("copy_mode", "missing_only").strip()

            allowed_copy_stores = allowed_store_numbers

            if not source_store or not destination_store:
                flash("Choose both a source store and destination store.", "error")
                return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

            if source_store == destination_store:
                flash("Source and destination stores must be different.", "error")
                return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

            if source_store not in allowed_copy_stores or destination_store not in allowed_copy_stores:
                flash("You do not have access to one of the selected stores.", "error")
                return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

            if copy_mode not in {"missing_only", "overwrite_matching"}:
                copy_mode = "missing_only"

            source_items = PrepTemplateItem.query.filter_by(
                store_number=source_store
            ).order_by(
                PrepTemplateItem.section_name.asc(),
                PrepTemplateItem.sort_order.asc(),
                PrepTemplateItem.id.asc()
            ).all()

            if not source_items:
                flash(f"No prep items found in source store {source_store}.", "error")
                return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

            destination_items = PrepTemplateItem.query.filter_by(
                store_number=destination_store
            ).all()

            destination_lookup = {
                normalize_item_key(item.item_name): item
                for item in destination_items
            }

            created_count = 0
            updated_count = 0
            skipped_count = 0

            for source_item in source_items:
                item_key = normalize_item_key(source_item.item_name)
                existing_item = destination_lookup.get(item_key)

                if existing_item:
                    if copy_mode == "overwrite_matching":
                        existing_item.section_name = source_item.section_name
                        existing_item.build_to = source_item.build_to
                        existing_item.instructions = source_item.instructions

                        existing_item.report_item_name = source_item.report_item_name
                        existing_item.prep_unit = source_item.prep_unit
                        existing_item.rounding_increment = source_item.rounding_increment
                        existing_item.minimum_build = source_item.minimum_build
                        existing_item.buffer_percent = source_item.buffer_percent
                        existing_item.conversion_notes = source_item.conversion_notes

                        existing_item.monday = source_item.monday
                        existing_item.tuesday = source_item.tuesday
                        existing_item.wednesday = source_item.wednesday
                        existing_item.thursday = source_item.thursday
                        existing_item.friday = source_item.friday
                        existing_item.saturday = source_item.saturday
                        existing_item.sunday = source_item.sunday
                        existing_item.sort_order = source_item.sort_order
                        existing_item.is_active = source_item.is_active
                        updated_count += 1
                    else:
                        skipped_count += 1
                    continue

                copied_item = PrepTemplateItem(
                    store_number=destination_store,
                    section_name=source_item.section_name,
                    item_name=source_item.item_name,
                    build_to=source_item.build_to,
                    instructions=source_item.instructions,
                    report_item_name=source_item.report_item_name,
                    prep_unit=source_item.prep_unit,
                    rounding_increment=source_item.rounding_increment,
                    minimum_build=source_item.minimum_build,
                    buffer_percent=source_item.buffer_percent,
                    conversion_notes=source_item.conversion_notes,
                    monday=source_item.monday,
                    tuesday=source_item.tuesday,
                    wednesday=source_item.wednesday,
                    thursday=source_item.thursday,
                    friday=source_item.friday,
                    saturday=source_item.saturday,
                    sunday=source_item.sunday,
                    sort_order=source_item.sort_order,
                    is_active=source_item.is_active,
                )
                db.session.add(copied_item)
                created_count += 1

            db.session.commit()

            if copy_mode == "overwrite_matching":
                flash(
                    f"Copied prep setup from {source_store} to {destination_store}. "
                    f"Created {created_count}, updated {updated_count}, skipped {skipped_count}.",
                    "success"
                )
            else:
                flash(
                    f"Copied missing prep items from {source_store} to {destination_store}. "
                    f"Created {created_count}, skipped {skipped_count} existing.",
                    "success"
                )

            return redirect(url_for("prep.manage", store=destination_store, show_inactive=int(show_inactive)))

        elif action == "delete":
            item_id = request.form.get("item_id", "").strip()
            item = PrepTemplateItem.query.get(item_id)

            if not item or item.store_number != store_number:
                flash("Prep item not found.", "error")
                return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

            usage_exists = DailyPrepItem.query.filter_by(template_item_id=item.id).first()

            if usage_exists:
                item.is_active = False
                db.session.commit()
                flash("Prep item archived because it already exists in prep history.", "success")
            else:
                db.session.delete(item)
                db.session.commit()
                flash("Prep item deleted.", "success")

            return redirect(url_for("prep.manage", store=store_number, show_inactive=int(show_inactive)))

    items_query = PrepTemplateItem.query.filter_by(
        store_number=store_number
    )

    if not show_inactive:
        items_query = items_query.filter_by(is_active=True)

    items = items_query.order_by(
        PrepTemplateItem.section_name.asc(),
        PrepTemplateItem.sort_order.asc(),
        PrepTemplateItem.id.asc()
    ).all()

    grouped_items = build_grouped_items(items)
    inactive_count = PrepTemplateItem.query.filter_by(store_number=store_number, is_active=False).count()

    return render_template(
        "prep/manage.html",
        items=items,
        grouped_items=grouped_items,
        stores=visible_stores,
        store_number=store_number,
        section_options=SECTION_OPTIONS,
        show_inactive=show_inactive,
        inactive_count=inactive_count,
        upload_preview=upload_preview,
    )